import os
from openpyxl import load_workbook,Workbook
names = ['01五年在監', '02入監罪名', '03入監刑名', '04入監教育', '05入監年齡', '06出獄罪名', '07假釋罪名', '08在監罪名', '09在監應執刑名', '10在監教育', '11在監年齡', '12戒治五年人數', 
'13戒治新入毒品級別', '14戒治新入教育', '15戒治新入年齡', '16戒治在所毒品級別', '17戒治在所教育', '18戒治在所年齡']

input_file="D:/外網統計園地/python/統計園地上網.xlsm"
output_folder="D:/外網統計園地/python"
def split_workbook_by_names(input_file, output_folder, names):
    """
    Split workbook into separate ods files based on names list.
    """

    workbook = load_workbook(input_file,read_only=True,data_only=True)
    print(workbook.sheetnames)
    
    for name in names:
        workbooksheet = workbook[name]
        # create new workbook for given name
        new_workbook = Workbook()
        new_workbook.remove(new_workbook.active)
        new_workbook.create_sheet(title=name)

        # copy rows from original workbook to new workbook
        for row in workbooksheet.iter_rows():
            #print(row[0].value)
            #if row[0].value == name:
            new_row = [cell.value for cell in row]
            print(new_row)
            new_workbook[name].append(new_row)

        # save new workbook as separate ods file
        new_file = os.path.join(output_folder, name + ".ods")
        new_workbook.save(new_file)

    workbook.close()
    new_workbook.close()

if __name__ == "__main__":
    input_file="D:/外網統計園地/python/統計園地上網.xlsm"
    output_folder="D:/外網統計園地/python"
    split_workbook_by_names(input_file, output_folder, names)