
import openpyxl

# 讀取Excel檔案
workbook = openpyxl.load_workbook('統計園地上網.xlsm',read_only=True)

# 取得所有sheet名稱
sheet_names = workbook.sheetnames

# 依照每個sheet名稱建立新的Excel檔案
for sheet_name in sheet_names:
    # 建立新的Excel檔案
    new_workbook = openpyxl.Workbook()
    # 取得sheet
    sheet = workbook[sheet_name]
    # 複製sheet到新的Excel檔案
    new_sheet = new_workbook.active
    for row in sheet.iter_rows():
        new_row = [cell.value for cell in row]
        print(new_row)
        new_sheet.append(new_row)
    # 儲存新的Excel檔案
    new_workbook.save(f'{sheet_name}.xlsx')

